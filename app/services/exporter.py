from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.config import EXPORTS_DIR
from app.core.models import AnalysisResult, Language


COLORS = {
    "navy": "07111F",
    "panel": "0B1728",
    "panel2": "10243A",
    "cyan": "2EE7D2",
    "blue": "3B82F6",
    "green": "26D07C",
    "amber": "F2B84B",
    "orange": "FF8A3D",
    "red": "FF4D6D",
    "ink": "E8F3FF",
    "muted": "91A6BD",
    "line": "24445E",
    "white": "FFFFFF",
}

SEVERITY_FILL = {
    "Critical": COLORS["red"],
    "High": COLORS["orange"],
    "Medium": COLORS["amber"],
    "Low": COLORS["green"],
}

LABELS = {
    "en": {
        "dashboard": "Dashboard",
        "iocs": "IOCs",
        "mitre": "MITRE Mapping",
        "actions": "Actions",
        "sources": "Source Details",
        "title": "CTI Intelligence Report",
        "subtitle": "IOC normalization, source agreement, MITRE ATT&CK mapping, and action-oriented reporting",
        "company": "Company",
        "date": "Analysis Date",
        "total": "Unique IOCs",
        "average": "Average Risk",
        "highCritical": "High / Critical",
        "highConfidence": "High Confidence",
        "submitted": "Submitted Values",
        "deduped": "Duplicates Removed",
        "executive": "Executive Summary",
        "operations": "Operations Summary",
        "technical": "Technical Summary",
        "severity": "Severity",
        "count": "Count",
        "type": "Type",
        "indicator": "Indicator",
        "risk": "Risk Score",
        "source": "Source",
        "status": "Status",
        "summary": "Summary",
        "tactic": "Tactic",
        "techniqueId": "Technique ID",
        "technique": "Technique",
        "confidence": "Confidence",
        "rationale": "Rationale",
        "action": "Recommended Action",
        "severityChart": "Severity Distribution",
        "typeChart": "IOC Types",
        "occurrences": "Occurrences",
        "normalizedFrom": "Normalized From",
        "agreement": "Source Agreement",
        "labels": "Threat Labels",
        "related": "Related Entities",
        "case": "Case",
        "caseStatus": "Case Status",
        "casePriority": "Case Priority",
        "caseCategory": "Case Category",
    },
    "ar": {
        "dashboard": "لوحة القيادة",
        "iocs": "المؤشرات",
        "mitre": "ربط MITRE",
        "actions": "الإجراءات",
        "sources": "تفاصيل المصادر",
        "title": "تقرير استخبارات التهديدات",
        "subtitle": "توحيد المؤشرات، اتفاق المصادر، ربط MITRE ATT&CK، وتوصيات قابلة للتنفيذ",
        "company": "الشركة",
        "date": "تاريخ التحليل",
        "total": "المؤشرات الفريدة",
        "average": "متوسط الخطورة",
        "highCritical": "عال / حرج",
        "highConfidence": "ثقة عالية",
        "submitted": "القيم المدخلة",
        "deduped": "التكرارات المحذوفة",
        "executive": "الملخص التنفيذي",
        "operations": "ملخص العمليات",
        "technical": "الملخص التقني",
        "severity": "التصنيف",
        "count": "العدد",
        "type": "النوع",
        "indicator": "المؤشر",
        "risk": "درجة الخطورة",
        "source": "المصدر",
        "status": "الحالة",
        "summary": "الملخص",
        "tactic": "Tactic",
        "techniqueId": "Technique ID",
        "technique": "Technique",
        "confidence": "الثقة",
        "rationale": "المبرر",
        "action": "الإجراء الموصى به",
        "severityChart": "توزيع الخطورة",
        "typeChart": "أنواع المؤشرات",
        "occurrences": "عدد التكرار",
        "normalizedFrom": "تم التوحيد من",
        "agreement": "اتفاق المصادر",
        "labels": "وسوم التهديد",
        "related": "كيانات مرتبطة",
        "case": "الحالة",
        "caseStatus": "حالة القضية",
        "casePriority": "أولوية القضية",
        "caseCategory": "تصنيف القضية",
    },
}


def tr(language: Language, key: str) -> str:
    return LABELS.get(language, LABELS["en"]).get(key, key)


def setup_sheet(ws, language: Language) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.rightToLeft = language == "ar"
    ws.freeze_panes = None
    ws.sheet_properties.tabColor = COLORS["cyan"]


def fill_range(ws, cell_range: str, color: str) -> None:
    fill = PatternFill("solid", fgColor=color)
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = fill


def set_border(ws, cell_range: str, color: str = "24445E") -> None:
    side = Side(style="thin", color=color)
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border


def style_table(ws, header_row: int, last_row: int, last_col: int, language: Language) -> None:
    for cell in ws[header_row]:
        if cell.column <= last_col:
            cell.fill = PatternFill("solid", fgColor=COLORS["panel2"])
            cell.font = Font(color=COLORS["cyan"], bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=header_row + 1, max_row=last_row, max_col=last_col):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=COLORS["panel"] if cell.row % 2 else "0E1D31")
            cell.font = Font(color=COLORS["ink"])
            cell.alignment = Alignment(horizontal="right" if language == "ar" else "left", vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="thin", color=COLORS["line"]))
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
    for col_idx in range(1, last_col + 1):
        max_len = max(len(str(ws.cell(row=row, column=col_idx).value or "")) for row in range(header_row, last_row + 1))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 3, 14), 52)


def add_title(ws, analysis: AnalysisResult, language: Language) -> None:
    ws.merge_cells("A1:H1")
    ws["A1"] = tr(language, "title")
    ws["A1"].font = Font(color=COLORS["white"], bold=True, size=22)
    ws["A1"].alignment = Alignment(horizontal="right" if language == "ar" else "left")
    ws.merge_cells("A2:H2")
    ws["A2"] = tr(language, "subtitle")
    ws["A2"].font = Font(color=COLORS["muted"], size=11)
    ws["A2"].alignment = Alignment(horizontal="right" if language == "ar" else "left")
    fill_range(ws, "A1:H2", COLORS["navy"])
    ws["A4"] = tr(language, "company")
    ws["B4"] = analysis.company_name
    ws["D4"] = tr(language, "date")
    ws["E4"] = analysis.created_at.strftime("%Y-%m-%d %H:%M UTC")
    for cell in ["A4", "D4"]:
        ws[cell].font = Font(color=COLORS["cyan"], bold=True)
    for cell in ["B4", "E4"]:
        ws[cell].font = Font(color=COLORS["ink"], bold=True)
    fill_range(ws, "A4:E4", COLORS["panel"])
    case = analysis.case or {}
    ws["A5"] = tr(language, "case")
    ws["B5"] = case.get("name") or case.get("status") or "No Case"
    ws["D5"] = tr(language, "casePriority")
    ws["E5"] = case.get("priority", "")
    for cell in ["A5", "D5"]:
        ws[cell].font = Font(color=COLORS["cyan"], bold=True)
    for cell in ["B5", "E5"]:
        ws[cell].font = Font(color=COLORS["ink"], bold=True)
    fill_range(ws, "A5:E5", COLORS["panel"])


def add_kpi(ws, cell: str, title: str, value: object, color: str) -> None:
    row = ws[cell].row
    col = ws[cell].column
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 1)
    ws.cell(row=row, column=col).value = title
    ws.cell(row=row + 1, column=col).value = value
    for r in range(row, row + 3):
        for c in range(col, col + 2):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=COLORS["panel2"])
            ws.cell(r, c).border = Border(
                left=Side(style="thin", color=color),
                right=Side(style="thin", color=color),
                top=Side(style="thin", color=color),
                bottom=Side(style="thin", color=color),
            )
            ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=col).font = Font(color=COLORS["muted"], bold=True)
    ws.cell(row=row + 1, column=col).font = Font(color=color, bold=True, size=24)


def build_dashboard(wb: Workbook, analysis: AnalysisResult, language: Language) -> None:
    ws = wb.active
    ws.title = tr(language, "dashboard")
    setup_sheet(ws, language)
    add_title(ws, analysis, language)
    high_critical = analysis.stats.get("by_severity", {}).get("High", 0) + analysis.stats.get("by_severity", {}).get("Critical", 0)
    add_kpi(ws, "A6", tr(language, "total"), analysis.stats.get("total", 0), COLORS["cyan"])
    add_kpi(ws, "C6", tr(language, "average"), analysis.stats.get("average_score", 0), COLORS["blue"])
    add_kpi(ws, "E6", tr(language, "highCritical"), high_critical, COLORS["orange"])
    add_kpi(ws, "G6", tr(language, "highConfidence"), analysis.stats.get("by_confidence", {}).get("High", 0), COLORS["green"])
    add_kpi(ws, "A10", tr(language, "submitted"), analysis.stats.get("submitted_count", analysis.stats.get("total", 0)), COLORS["muted"])
    add_kpi(ws, "C10", tr(language, "deduped"), analysis.stats.get("deduplicated_count", 0), COLORS["amber"])

    summaries = [
        (tr(language, "executive"), analysis.summaries.executive),
        (tr(language, "operations"), analysis.summaries.operations),
        (tr(language, "technical"), analysis.summaries.technical),
    ]
    row = 15
    for title, text in summaries:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row=row, column=1).value = title
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=COLORS["panel2"])
        ws.cell(row=row, column=1).font = Font(color=COLORS["cyan"], bold=True, size=12)
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 3, end_column=8)
        ws.cell(row=row + 1, column=1).value = text
        ws.cell(row=row + 1, column=1).fill = PatternFill("solid", fgColor=COLORS["panel"])
        ws.cell(row=row + 1, column=1).font = Font(color=COLORS["ink"])
        ws.cell(row=row + 1, column=1).alignment = Alignment(horizontal="right" if language == "ar" else "left", vertical="top", wrap_text=True)
        set_border(ws, f"A{row}:H{row + 3}")
        row += 5

    severity_start = 32
    ws.cell(severity_start, 1).value = tr(language, "severity")
    ws.cell(severity_start, 2).value = tr(language, "count")
    for idx, (severity, count) in enumerate(analysis.stats.get("by_severity", {}).items(), start=severity_start + 1):
        ws.cell(idx, 1).value = severity
        ws.cell(idx, 2).value = count
    type_start = 32
    ws.cell(type_start, 4).value = tr(language, "type")
    ws.cell(type_start, 5).value = tr(language, "count")
    type_items = [(k, v) for k, v in analysis.stats.get("by_type", {}).items() if v]
    for idx, (kind, count) in enumerate(type_items, start=type_start + 1):
        ws.cell(idx, 4).value = kind
        ws.cell(idx, 5).value = count
    style_table(ws, severity_start, severity_start + 4, 2, language)
    if type_items:
        style_table(ws, type_start, type_start + len(type_items), 5, language)

    pie = DoughnutChart()
    pie.title = tr(language, "severityChart")
    labels = Reference(ws, min_col=1, min_row=severity_start + 1, max_row=severity_start + 4)
    data = Reference(ws, min_col=2, min_row=severity_start, max_row=severity_start + 4)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.holeSize = 55
    ws.add_chart(pie, "G32")

    if type_items:
        bar = BarChart()
        bar.title = tr(language, "typeChart")
        labels = Reference(ws, min_col=4, min_row=type_start + 1, max_row=type_start + len(type_items))
        data = Reference(ws, min_col=5, min_row=type_start, max_row=type_start + len(type_items))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(labels)
        bar.style = 10
        ws.add_chart(bar, "G48")

    labels_start = 42
    ws.cell(labels_start, 1).value = tr(language, "labels")
    ws.cell(labels_start, 2).value = tr(language, "count")
    label_items = sorted(analysis.stats.get("threat_labels", {}).items(), key=lambda item: item[1], reverse=True)[:8]
    for idx, (label, count) in enumerate(label_items, start=labels_start + 1):
        ws.cell(idx, 1).value = label
        ws.cell(idx, 2).value = count
    if label_items:
        style_table(ws, labels_start, labels_start + len(label_items), 2, language)

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18
    for row_idx in range(1, 65):
        ws.row_dimensions[row_idx].height = 22


def agreement_text(item) -> str:
    text = f"M={item.source_agreement.get('malicious', 0)}, S={item.source_agreement.get('suspicious', 0)}, C={item.source_agreement.get('clean', 0)}"
    unavailable = item.source_agreement.get("source_unavailable", 0)
    if unavailable:
        text += f", U={unavailable}"
    return text


def build_iocs(wb: Workbook, analysis: AnalysisResult, language: Language) -> None:
    ws = wb.create_sheet(tr(language, "iocs"))
    setup_sheet(ws, language)
    headers = [
        tr(language, "indicator"),
        tr(language, "type"),
        tr(language, "risk"),
        tr(language, "severity"),
        tr(language, "confidence"),
        tr(language, "agreement"),
        tr(language, "labels"),
        tr(language, "occurrences"),
        tr(language, "normalizedFrom"),
        tr(language, "related"),
        "MITRE",
        tr(language, "summary"),
    ]
    ws.append(headers)
    for item in analysis.indicators:
        ws.append([
            item.value,
            item.type,
            item.risk_score,
            item.severity,
            item.confidence,
            agreement_text(item),
            ", ".join(item.threat_labels),
            item.occurrence_count,
            " | ".join(item.normalized_from),
            " | ".join(f"{k}={v}" for k, v in item.related_entities.items()),
            " | ".join(f"{m.technique_id} {m.technique}" for m in item.mitre),
            " | ".join(f"{v.source}: {v.status} ({v.summary})" for v in item.verdicts),
        ])
    style_table(ws, 1, max(ws.max_row, 1), len(headers), language)
    for row in range(2, ws.max_row + 1):
        severity = ws.cell(row=row, column=4).value
        color = SEVERITY_FILL.get(severity, COLORS["muted"])
        ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor=color)
        ws.cell(row=row, column=4).font = Font(color=COLORS["navy"], bold=True)
        confidence = ws.cell(row=row, column=5).value
        confidence_color = {"High": COLORS["cyan"], "Medium": COLORS["blue"], "Low": COLORS["muted"]}.get(confidence, COLORS["muted"])
        ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor=confidence_color)
        ws.cell(row=row, column=5).font = Font(color=COLORS["navy"] if confidence == "High" else COLORS["white"], bold=True)


def build_mitre(wb: Workbook, analysis: AnalysisResult, language: Language) -> None:
    ws = wb.create_sheet(tr(language, "mitre"))
    setup_sheet(ws, language)
    headers = [tr(language, "indicator"), tr(language, "confidence"), tr(language, "labels"), tr(language, "tactic"), tr(language, "techniqueId"), tr(language, "technique"), tr(language, "rationale")]
    ws.append(headers)
    for item in analysis.indicators:
        for mitre in item.mitre:
            ws.append([item.value, item.confidence, ", ".join(item.threat_labels), mitre.tactic, mitre.technique_id, mitre.technique, mitre.rationale])
    style_table(ws, 1, max(ws.max_row, 1), len(headers), language)


def build_actions(wb: Workbook, analysis: AnalysisResult, language: Language) -> None:
    ws = wb.create_sheet(tr(language, "actions"))
    setup_sheet(ws, language)
    headers = [tr(language, "indicator"), tr(language, "severity"), tr(language, "confidence"), tr(language, "labels"), tr(language, "action")]
    ws.append(headers)
    for item in analysis.indicators:
        for action in item.recommended_actions:
            ws.append([item.value, item.severity, item.confidence, ", ".join(item.threat_labels), action])
    style_table(ws, 1, max(ws.max_row, 1), len(headers), language)


def build_sources(wb: Workbook, analysis: AnalysisResult, language: Language) -> None:
    ws = wb.create_sheet(tr(language, "sources"))
    setup_sheet(ws, language)
    headers = [tr(language, "indicator"), tr(language, "confidence"), tr(language, "source"), tr(language, "status"), tr(language, "risk"), tr(language, "summary")]
    ws.append(headers)
    for item in analysis.indicators:
        for verdict in item.verdicts:
            ws.append([item.value, item.confidence, verdict.source, verdict.status, verdict.score, verdict.summary])
    style_table(ws, 1, max(ws.max_row, 1), len(headers), language)


def export_excel(analysis: AnalysisResult) -> Path:
    language: Language = analysis.language or "ar"
    wb = Workbook()
    build_dashboard(wb, analysis, language)
    build_iocs(wb, analysis, language)
    build_mitre(wb, analysis, language)
    build_actions(wb, analysis, language)
    build_sources(wb, analysis, language)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                horizontal = cell.alignment.horizontal if cell.alignment.horizontal == "center" else ("right" if language == "ar" else "left")
                cell.alignment = Alignment(
                    horizontal=horizontal,
                    vertical=cell.alignment.vertical or "top",
                    wrap_text=True,
                )
        ws.sheet_view.rightToLeft = language == "ar"

    path = EXPORTS_DIR / f"cti_report_{analysis.id}_{language}.xlsx"
    wb.save(path)
    return path


def build_case_overview(wb: Workbook, case: dict, language: Language) -> None:
    ws = wb.active
    ws.title = "Case Overview"
    setup_sheet(ws, language)
    ws.merge_cells("A1:H1")
    ws["A1"] = "CTI Case Report"
    ws["A1"].font = Font(color=COLORS["white"], bold=True, size=22)
    ws.merge_cells("A2:H2")
    ws["A2"] = case.get("name", case.get("id", "Case"))
    ws["A2"].font = Font(color=COLORS["muted"], size=12)
    fill_range(ws, "A1:H2", COLORS["navy"])

    rows = [
        ("Case ID", case.get("id", "")),
        ("Status", case.get("status", "")),
        ("Priority", case.get("priority", "")),
        ("Category", case.get("category", "")),
        ("Created", case.get("created_at", "")),
        ("Updated", case.get("updated_at", "")),
        ("Auto Created", "Yes" if case.get("auto_created") else "No"),
        ("Rationale", case.get("reason", "")),
        ("Analyst Notes", case.get("notes", "")),
    ]
    start = 4
    for idx, (label, value) in enumerate(rows, start=start):
        ws.cell(idx, 1).value = label
        ws.cell(idx, 2).value = value
    style_table(ws, start, start + len(rows) - 1, 2, language)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90


def build_case_tasks(wb: Workbook, case: dict, language: Language) -> None:
    ws = wb.create_sheet("Checklist")
    setup_sheet(ws, language)
    headers = ["Task", "Completed", "Updated"]
    ws.append(headers)
    for task in case.get("tasks", []):
        ws.append([task.get("title", ""), "Yes" if task.get("completed") else "No", task.get("updated_at", "")])
    style_table(ws, 1, max(ws.max_row, 1), len(headers), language)


def build_case_timeline(wb: Workbook, case: dict, language: Language) -> None:
    ws = wb.create_sheet("Timeline")
    setup_sheet(ws, language)
    headers = ["Time", "Event", "Message"]
    ws.append(headers)
    for event in case.get("timeline", []):
        ws.append([event.get("created_at", ""), event.get("event_type", ""), event.get("message", "")])
    style_table(ws, 1, max(ws.max_row, 1), len(headers), language)


def build_case_analyses(wb: Workbook, case: dict, language: Language) -> None:
    ws = wb.create_sheet("Analyses")
    setup_sheet(ws, language)
    headers = ["Analysis ID", "Created", "Average Risk", "Unique IOCs", "High/Critical", "High Confidence"]
    ws.append(headers)
    for analysis in case.get("analyses", []):
        ws.append([
            analysis.get("id", ""),
            analysis.get("created_at", ""),
            analysis.get("average_score", 0),
            analysis.get("total_iocs", 0),
            analysis.get("high_critical", 0),
            analysis.get("high_confidence", 0),
        ])
    style_table(ws, 1, max(ws.max_row, 1), len(headers), language)


def build_case_indicators(wb: Workbook, case: dict, language: Language) -> None:
    ws = wb.create_sheet("Indicators")
    setup_sheet(ws, language)
    headers = ["Indicator", "Type", "Severity", "Confidence", "Risk Score", "Threat Labels"]
    ws.append(headers)
    for indicator in case.get("indicators", []):
        ws.append([
            indicator.get("value", ""),
            indicator.get("type", ""),
            indicator.get("severity", ""),
            indicator.get("confidence", ""),
            indicator.get("risk_score", 0),
            indicator.get("threat_labels", ""),
        ])
    style_table(ws, 1, max(ws.max_row, 1), len(headers), language)


def export_case_excel(case: dict, language: Language = "en") -> Path:
    wb = Workbook()
    build_case_overview(wb, case, language)
    build_case_tasks(wb, case, language)
    build_case_timeline(wb, case, language)
    build_case_analyses(wb, case, language)
    build_case_indicators(wb, case, language)
    for ws in wb.worksheets:
        ws.sheet_view.rightToLeft = language == "ar"
    path = EXPORTS_DIR / f"cti_case_report_{case.get('id', 'case')}_{language}.xlsx"
    wb.save(path)
    return path
